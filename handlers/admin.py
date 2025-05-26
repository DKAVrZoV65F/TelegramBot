# handlers/admin.py

from aiogram import types
from config import settings
from handlers.statistics import cmd_statistics, process_back_to_type_choice, process_statistics_period, process_statistics_request, process_statistics_type_choice
from keyboards import admin_menu, period_menu
from utils import make_msg_link
from db import get_conn
from datetime import datetime, timedelta
import io, xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

EMOJI = {"positive": "😃", "negative": "😡", "neutral": "😐"}


def _period_to_dt(code: str) -> datetime:
  now = datetime.utcnow()
  return {"day": now - timedelta(days=1),
      "week": now - timedelta(days=7),
      "month": now - timedelta(days=30)}[code]

async def show_menu(m: types.Message):
  if m.from_user.id not in settings.ADMIN_IDS:
    return await m.reply("⛔")

  await m.answer("📂 Главное меню:", reply_markup=admin_menu())

async def ask_period(call: types.CallbackQuery):
  await call.message.edit_reply_markup(period_menu("u:"))

async def ask_period_xml(call: types.CallbackQuery):
  await call.message.edit_reply_markup(period_menu("x:"))

async def _fetch_unprocessed(since: datetime):
  sql = """
    SELECT
      m.id,
      m.chat_id,
      COALESCE(LIST(t.name, ','), '')     AS tags,   -- 🔹 Firebird LIST()
      COALESCE(m.sentiment, 'neutral')    AS sentiment,
      COALESCE(m.text, '')          AS text
    FROM messages m
    LEFT JOIN message_tags mt
       ON m.id = mt.message_id AND m.chat_id = mt.chat_id
    LEFT JOIN tags t ON t.id = mt.tag_id
    WHERE (mt.processed IS NULL OR mt.processed = 0)
     AND m.msg_date >= ?
    GROUP BY m.id, m.chat_id, m.sentiment, m.text
    ORDER BY MAX(m.msg_date) DESC
  """

  con = get_conn()
  cur = con.cursor()

  cur.execute(sql, (since,))
  rows = cur.fetchall(); con.close()

  return rows

async def send_unprocessed(call: types.CallbackQuery):
  per  = call.data.split(":")[1]
  since = _period_to_dt(per)
  rows = await _fetch_unprocessed(since)

  if not rows:
    return await call.answer("🎉 Всё обработано!")

  sent = 0
  for mid, cid, raw_tags, sentiment, txt in rows[:50]:
    tags = [t for t in raw_tags.split(",") if t] or ["нет_тега"]
    link = make_msg_link(cid, mid)
    emoji = EMOJI.get(sentiment, "😐")
    prev = _html_escape((txt or "⋯").strip())[:400]

    body = (
      f"{emoji} <b>Новый инцидент</b> · <i>{sentiment.title()}</i>\n"
      f"{_format_badges(tags)}\n"
      f"<blockquote>{prev}</blockquote>\n"
      f"🔗 <a href='{link}'>Открыть сообщение</a>"
    )

    first_tag = tags[0]
    kb = types.InlineKeyboardMarkup(row_width=3)
    kb.add(
      types.InlineKeyboardButton("✅ Обработано", callback_data=f"mark:1:{cid}:{mid}:{first_tag}"),
      types.InlineKeyboardButton("❌ Отклонить", callback_data=f"mark:0:{cid}:{mid}:{first_tag}"),
      types.InlineKeyboardButton("🗑 Закрыть",  callback_data="close_notify"),
    )

    await call.bot.send_message(
      call.from_user.id,
      body,
      parse_mode="HTML",
      reply_markup=kb,
      disable_web_page_preview=True
    )
    sent += 1

  await call.answer(f"✅ Отправлено {sent}") 

async def send_unprocessed_xml(call: types.CallbackQuery):
  per = call.data.split(":")[1]
  since = _period_to_dt(per)

  sql = """
  SELECT
   m.id, m.chat_id, m.msg_date, m.user_id, m.username,
   COALESCE(m.text, '') as text,
   COALESCE(m.raw_json, '') as raw_json,
   COALESCE(LIST(t.name, ','), '') AS tags,
   COALESCE(m.sentiment, 'neutral') AS sentiment,
   COALESCE(mt.processed, 0) AS processed
  FROM messages m
  LEFT JOIN message_tags mt
   ON m.id=mt.message_id AND m.chat_id=mt.chat_id
  LEFT JOIN tags t
   ON t.id=mt.tag_id
  WHERE m.msg_date >= ?
  GROUP BY m.id, m.chat_id, m.msg_date, m.user_id, m.username,
      m.text, m.raw_json, m.sentiment, mt.processed
  ORDER BY m.msg_date DESC
  """

  con = get_conn()
  cur = con.cursor()

  cur.execute(sql, (since, ))
  rows = cur.fetchall()
  con.close()

  root = ET.Element("messages", {
    "period": per,
    "generated_at": datetime.utcnow().isoformat() + "Z",
    "count": str(len(rows))
  })

  for msg_id, chat_id, dt, user_id, username, text, raw_json, tags, sentiment, processed in rows:
    m = ET.SubElement(root, "message", {
      "id":    str(msg_id),
      "chat_id": str(chat_id),
      "msg_date": dt.isoformat(),
      "user_id": str(user_id or ""),
      "username": username or "",
      "sentiment": sentiment,
      "processed": str(processed),
      "link":   make_msg_link(chat_id, msg_id)
    })

    tags_el = ET.SubElement(m, "tags")
    for tag in tags.split(","):
      if tag:
        ET.SubElement(tags_el, "tag", {"name": tag})

    text_el = ET.SubElement(m, "text")
    text_el.text = f"<![CDATA[{text}]]>"
    
    if raw_json:
      raw_el = ET.SubElement(m, "raw_json")
      raw_el.text = f"<![CDATA[{raw_json}]]>"

  buf = io.BytesIO()
  ET.ElementTree(root).write(buf, encoding="utf-8", xml_declaration=True)
  buf.seek(0)

  await call.message.answer_document(
    ("report.xml", buf.read()),
    caption=f"XML · {len(rows)} записей (период: {per})"
  )

  await call.answer()

async def export_xlsx(call: types.CallbackQuery):
  since = datetime.utcnow() - timedelta(days=365)
  
  sql = """
  SELECT
   m.id, m.chat_id, m.msg_date, m.user_id, m.username,
   COALESCE(m.text, '') AS text,
   COALESCE(m.raw_json, '') AS raw_json,
   COALESCE(LIST(t.name, ','), '') AS tags,
   COALESCE(m.sentiment, '') AS sentiment,
   COALESCE(mt.processed, 0) AS processed
  FROM messages m
  LEFT JOIN message_tags mt
   ON m.id=mt.message_id AND m.chat_id=mt.chat_id
  LEFT JOIN tags t
   ON t.id=mt.tag_id
  WHERE m.msg_date >= ?
  GROUP BY m.id, m.chat_id, m.msg_date, m.user_id, m.username,
      m.text, m.raw_json, m.sentiment, mt.processed
  ORDER BY m.msg_date DESC
  """

  con = get_conn()
  cur = con.cursor()

  cur.execute(sql, (since,))
  rows = cur.fetchall()
  con.close()

  if not rows:
    return await call.answer("Пусто 🙌")

  wb = Workbook()
  ws = wb.active
  ws.title = "Messages"

  headers = [
    "msg_id", "chat_id", "msg_date_utc", "user_id", "username",
    "text", "raw_json", "tags", "sentiment", "processed", "link"
  ]

  ws.append(headers)

  for msg_id, chat_id, dt, user_id, username, text, raw_json, tags, sentiment, processed in rows:
    link = make_msg_link(chat_id, msg_id)
    ws.append([
      msg_id, chat_id, dt.isoformat(),
      user_id or "", username or "",
      text.replace("\n", " ").strip(),
      raw_json.replace("\n", " ").strip(),
      tags, sentiment, processed, link
    ])

  for i, _ in enumerate(headers, start=1):
    col_letter = get_column_letter(i)
    ws.column_dimensions[col_letter].width = 20


  buf = io.BytesIO()
  wb.save(buf)
  buf.seek(0)

  await call.message.answer_document(
    ("report.xlsx", buf.read()),
    caption=f"XLSX · {len(rows)} строк (365d)"
  )

  await call.answer()

async def close_notify(call: types.CallbackQuery):
  try:
    await call.message.delete()
  finally:
    await call.answer()

def register(dp):
  dp.register_message_handler(show_menu, commands=["menu"])
  dp.register_callback_query_handler(ask_period,   lambda c: c.data == "req:uproc")
  dp.register_callback_query_handler(ask_period_xml, lambda c: c.data == "req:uproc_xml")
  dp.register_callback_query_handler(send_unprocessed,   lambda c: c.data.startswith("u:"))
  dp.register_callback_query_handler(send_unprocessed_xml, lambda c: c.data.startswith("x:"))
  dp.register_callback_query_handler(close_notify, lambda c: c.data == "close_notify")
  dp.register_callback_query_handler(ask_export_xlsx, lambda c: c.data == "req:export_xlsx")
  dp.register_callback_query_handler(export_xlsx, lambda c: c.data == "export_xlsx")
  dp.register_callback_query_handler(send_export_xlsx, lambda c: c.data.startswith("c:"))
  dp.register_message_handler(cmd_statistics, commands=["statistics"])
  dp.register_callback_query_handler(process_statistics_type_choice, lambda c: c.data.startswith("stats_type:"))
  dp.register_callback_query_handler(process_statistics_request, lambda c: c.data.startswith("stats_fetch:"))
  dp.register_callback_query_handler(process_back_to_type_choice, lambda c: c.data == "stats_back_to_type")
  dp.register_callback_query_handler(process_statistics_period, lambda c: c.data.startswith("stats_fetch:"))

def _format_badges(tags: list[str]) -> str:
  pad = lambda t: (t + " " * (10 - len(t)))[:10]
  return " ".join(f"<code>#{pad(t)}</code>" for t in tags)
  
def _html_escape(text: str) -> str:
  return (text or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

async def ask_export_xlsx(call: types.CallbackQuery):
  await call.message.edit_text("Выберите период для XLSX-экспорта:",
                reply_markup=period_menu("c:"))
  await call.answer()

async def send_export_xlsx(call: types.CallbackQuery):
  _, period = call.data.split(":")
  from_date = _period_to_dt(period)

  sql = """
  SELECT
   m.id, m.chat_id, m.msg_date, m.user_id, m.username,
   COALESCE(m.text,'')   AS text,
   COALESCE(m.raw_json,'') AS raw_json,
   COALESCE(LIST(t.name,','), '') AS tags,
   COALESCE(m.sentiment,'') AS sentiment
  FROM messages m
  LEFT JOIN message_tags mt
   ON m.id=mt.message_id AND m.chat_id=mt.chat_id
  LEFT JOIN tags t
   ON t.id=mt.tag_id
  WHERE m.msg_date >= ?
  GROUP BY m.id,m.chat_id,m.msg_date,m.user_id,m.username,
      m.text,m.raw_json,m.sentiment
  ORDER BY m.msg_date DESC
  """

  con = get_conn()
  cur = con.cursor()

  cur.execute(sql, (from_date,))
  rows = cur.fetchall()
  con.close()

  if not rows:
    return await call.answer("Ничего нет за этот период.")

  wb = Workbook()
  ws = wb.active
  ws.title = "Export"

  headers = [
    "msg_id","chat_id","msg_date_utc","user_id","username",
    "text","raw_json","tags","sentiment","link"
  ]

  ws.append(headers)

  for mid, cid, dt, uid, uname, text, raw, tags, sent in rows:
    ws.append([
      mid, cid, dt.isoformat(),
      uid or "", uname or "",
      text.replace("\n"," "),
      raw.replace("\n"," "),
      tags, 
      sent,
      make_msg_link(cid, mid)
    ])

  for i, _ in enumerate(headers, start=1):
    ws.column_dimensions[get_column_letter(i)].width = 20

  buf = io.BytesIO()
  wb.save(buf)
  buf.seek(0)

  await call.message.answer_document(
    ("export.xlsx", buf.read()),
    caption=f"XLSX-экспорт: {period} • {len(rows)} строк"
  )

  await call.answer()
